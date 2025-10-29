import os
import sys
import subprocess
import datetime as dt
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill

# -----------------------
# CONFIGURATION
# -----------------------
FILE_NAME = "MyDailyLifeRepert.xlsx"

SHEET_TRACKER = "Tracker"
SHEET_REPORTS = "Reports"        # Totals per activity + chart
SHEET_SUMMARY = "Summary"        # Averages + Overall/D/W/M + chart
SHEET_EXPORT  = "Report_Export"  # Printable Daily/Weekly/Monthly tables

def ensure_sheets_exist():
    """Make sure Tracker, Reports, and Summary sheets exist with proper headers."""
    if not os.path.exists(FILE_NAME):
        # Create a blank workbook
        wb = openpyxl.Workbook()
        # Tracker sheet
        ws = wb.active
        ws.title = SHEET_TRACKER
        ws.append(["Date", "Task", "Hours", "Mood"])
        # Reports sheet
        ws_reports = wb.create_sheet(SHEET_REPORTS)
        ws_reports.append(["Activity", "Hours"])
        for act in ACTIVITIES:
            ws_reports.append([act, 0])
        # Summary sheet
        ws_summary = wb.create_sheet(SHEET_SUMMARY)
        ws_summary.append(["Activity", "Average Hours"])
        for act in ACTIVITIES:
            ws_summary.append([act, 0])
        # Extra rows for overall/daily/weekly/monthly
        ws_summary.append(["Overall Average", 0])
        ws_summary.append(["Daily Average", 0])
        ws_summary.append(["Weekly Average", 0])
        ws_summary.append(["Monthly Average", 0])

        wb.save(FILE_NAME)

ACTIVITIES = [
    "EXERCISE", "MEDITATION", "COOKING", "READING", "STUDY",
    "WORK", "COLLEGE", "TUTION", "LEISURE", "OTHER"
]

# Totals sheet fixed cell map (Reports)
CELL_MAP = {
    "EXERCISE": "B2", "MEDITATION": "B3", "COOKING": "B4",
    "READING": "B5", "STUDY": "B6", "WORK": "B7",
    "COLLEGE": "B8", "TUTION": "B9", "LEISURE": "B10",
    "OTHER": "B11"
}

# Summary (Averages) fixed cell map
CELL_MAP_SUMMARY = {
    "EXERCISE": "B2",
    "MEDITATION": "B3",
    "COOKING": "B4",
    "READING": "B5",
    "STUDY": "B6",
    "WORK": "B7",
    "COLLEGE": "B8",
    "TUTION": "B9",
    "LEISURE": "B10",
    "OTHER": "B11",
    "OVERALL_AVG": "B13",
    "DAILY_AVG":   "B14",
    "WEEKLY_AVG":  "B15",
    "MONTHLY_AVG": "B16"
}

# -----------------------
# HELPERS
# -----------------------
def parse_hours(text: str) -> float:
    """Parse hours text like '1 hr 30 min' or '1.5 hr' into float hours."""
    text = text.strip().lower()
    hours = 0.0
    parts = text.split()
    i = 0
    while i < len(parts):
        try:
            val = float(parts[i])
            if i + 1 < len(parts):
                unit = parts[i + 1]
                if "hr" in unit or "hour" in unit:
                    hours += val
                    i += 2
                    continue
                elif "min" in unit:
                    hours += val / 60
                    i += 2
                    continue
            hours += val
            i += 1
        except ValueError:
            i += 1
    if hours == 0.0:
        try:
            hours = float(text)
        except ValueError:
            hours = 0.0
    return hours

def get_mood() -> str:
    """Prompt user to choose a mood from fixed options."""
    moods = ["Happy", "Sad", "Neutral"]
    print("\nSelect mood:")
    for i, mood in enumerate(moods, start=1):
        print(f"{i}. {mood}")
    try:
        choice = int(input("Enter mood number: ").strip())
        if 1 <= choice <= len(moods):
            return moods[choice - 1]
    except ValueError:
        pass
    print("⚠ Invalid mood. Defaulting to Neutral.")
    return "Neutral"

# -----------------------
# REPORTS (TOTALS) SHEET
# -----------------------
def setup_reports_sheet():
    """Create Reports sheet with formatting if not exists."""
    wb = load_workbook(FILE_NAME)
    if SHEET_REPORTS not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_REPORTS)
        ws["A1"] = "Activity"
        ws["B1"] = "Hours"
        for i, act in enumerate(ACTIVITIES, start=2):
            ws[f"A{i}"] = act
            ws[f"B{i}"] = 0
    else:
        ws = wb[SHEET_REPORTS]

    # Styles
    header_font = Font(name="Cambria", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    activity_font = Font(name="Cambria", size=11, color="000000")
    activity_align = Alignment(horizontal="center", vertical="center")

    ws["A1"].font = ws["B1"].font = header_font
    ws["A1"].fill = ws["B1"].fill = header_fill
    ws["A1"].alignment = ws["B1"].alignment = header_align

    for i in range(2, len(ACTIVITIES) + 2):
        ws[f"A{i}"].font = ws[f"B{i}"].font = activity_font
        ws[f"A{i}"].alignment = ws[f"B{i}"].alignment = activity_align

    wb.save(FILE_NAME)

def add_to_reports(activity, hours):
    """Update Reports sheet totals and chart."""
    wb = load_workbook(FILE_NAME)
    ws = wb[SHEET_REPORTS]

    act_upper = activity.upper()
    if act_upper in CELL_MAP:
        cell = CELL_MAP[act_upper]
        current_value = ws[cell].value or 0
        ws[cell] = float(current_value) + hours

    # Chart (totals)
    chart = BarChart()
    chart.title = "Total Hours by Activity"
    chart.x_axis.title = "Activity"
    chart.y_axis.title = "Hours"

    data = Reference(ws, min_col=2, min_row=1, max_row=len(ACTIVITIES) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(ACTIVITIES) + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws._charts = []  # clear previous
    ws.add_chart(chart, "D2")

    wb.save(FILE_NAME)

# -----------------------
# SUMMARY (AVERAGES) SHEET
# -----------------------
def setup_summary_sheet():
    """Create Summary sheet for averages if not exists, with headings and extra rows."""
    wb = load_workbook(FILE_NAME)
    if SHEET_SUMMARY not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_SUMMARY)

        # Headers
        ws["A1"] = "Activity"
        ws["B1"] = "Average Hours"

        # Activity rows
        for i, act in enumerate(ACTIVITIES, start=2):
            ws[f"A{i}"] = act
            ws[f"B{i}"] = 0

        # Overall + time-based averages
        ws["A13"] = "Overall Average"; ws["B13"] = 0
        ws["A14"] = "Daily Average";   ws["B14"] = 0
        ws["A15"] = "Weekly Average";  ws["B15"] = 0
        ws["A16"] = "Monthly Average"; ws["B16"] = 0
    else:
        ws = wb[SHEET_SUMMARY]

    # Styles
    header_font = Font(name="Cambria", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    activity_font = Font(name="Cambria", size=11, color="000000")
    activity_align = Alignment(horizontal="center", vertical="center")

    ws["A1"].font = ws["B1"].font = header_font
    ws["A1"].fill = ws["B1"].fill = header_fill
    ws["A1"].alignment = ws["B1"].alignment = header_align

    for i in range(2, 17):
        ws[f"A{i}"].font = ws[f"B{i}"].font = activity_font
        ws[f"A{i}"].alignment = ws[f"B{i}"].alignment = activity_align

    wb.save(FILE_NAME)

def update_summary_sheet():
    """Compute per-activity averages + Overall/Daily/Weekly/Monthly and update chart."""
    if not os.path.exists(FILE_NAME):
        return
    try:
        df = pd.read_excel(FILE_NAME, sheet_name=SHEET_TRACKER)
    except Exception:
        return
    if df.empty:
        return

    df["Date"] = pd.to_datetime(df["Date"])
    # Per-task average
    avg_per_task = df.groupby("Task")["Hours"].mean().round(2)

    # Overall average
    overall_avg = df["Hours"].mean().round(2)

    # Daily average
    daily_avg = df.groupby(df["Date"].dt.date)["Hours"].sum().mean().round(2)

    # Weekly average
    weekly = df.groupby(df["Date"].dt.to_period("W"))["Hours"].sum()
    weekly_avg = weekly.mean().round(2) if not weekly.empty else 0.0

    # Monthly average
    monthly = df.groupby(df["Date"].dt.to_period("M"))["Hours"].sum()
    monthly_avg = monthly.mean().round(2) if not monthly.empty else 0.0

    wb = load_workbook(FILE_NAME)
    ws = wb[SHEET_SUMMARY]

    # Fill averages
    for act, cell in CELL_MAP_SUMMARY.items():
        if act == "OVERALL_AVG":
            ws[cell] = overall_avg
        elif act == "DAILY_AVG":
            ws[cell] = daily_avg
        elif act == "WEEKLY_AVG":
            ws[cell] = weekly_avg
        elif act == "MONTHLY_AVG":
            ws[cell] = monthly_avg
        else:
            ws[cell] = float(avg_per_task.get(act, 0))

    # Chart (averages)
    chart = BarChart()
    chart.title = "Average Hours by Activity"
    chart.x_axis.title = "Activity"
    chart.y_axis.title = "Avg Hours"

    data = Reference(ws, min_col=2, min_row=1, max_row=len(ACTIVITIES) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(ACTIVITIES) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws._charts = []
    ws.add_chart(chart, "D2")

    wb.save(FILE_NAME)
    print("✅ Summary sheet updated with averages!")

# -----------------------
# REPORT EXPORT (DAILY/WEEKLY/MONTHLY)
# -----------------------
def setup_export_sheet():
    """Ensure Report_Export sheet exists with proper headers."""
    wb = load_workbook(FILE_NAME)
    if SHEET_EXPORT not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_EXPORT)
        ws.append(["Date/Period", "Activity", "Total Hours"])
    wb.save(FILE_NAME)


def export_report(period="daily"):
    """Generate and export a report by day, week, or month."""
    if not os.path.exists(FILE_NAME):
        print("⚠ No data to export.")
        return

    try:
        df = pd.read_excel(FILE_NAME, sheet_name=SHEET_TRACKER)
    except Exception as e:
        print("⚠ Error reading tracker sheet:", e)
        return

    if df.empty:
        print("⚠ Tracker is empty.")
        return

    df["Date"] = pd.to_datetime(df["Date"])

    if period == "daily":
        # Ask for a date
        date_str = input("Enter date (YYYY-MM-DD): ").strip()
        try:
            date = dt.datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            print("⚠ Invalid date.")
            return

        df_day = df[df["Date"].dt.date == date]
        if df_day.empty:
            print("⚠ No entries for that day.")
            return
        summary = df_day.groupby("Task")["Hours"].sum().reset_index()
        summary.insert(0, "Date/Period", str(date))

    elif period == "weekly":
        df["Week"] = df["Date"].dt.to_period("W").apply(lambda r: str(r.start_time.date()))
        summary = df.groupby(["Week", "Task"])["Hours"].sum().reset_index()
        summary.rename(columns={"Week": "Date/Period"}, inplace=True)

    elif period == "monthly":
        df["Month"] = df["Date"].dt.to_period("M").apply(lambda r: str(r.start_time.date())[:7])
        summary = df.groupby(["Month", "Task"])["Hours"].sum().reset_index()
        summary.rename(columns={"Month": "Date/Period"}, inplace=True)
    else:
        print("⚠ Invalid period type.")
        return

    wb = load_workbook(FILE_NAME)
    if SHEET_EXPORT not in wb.sheetnames:
        setup_export_sheet()
        wb = load_workbook(FILE_NAME)

    ws = wb[SHEET_EXPORT]
    ws.delete_rows(2, ws.max_row)  # clear old data
    for row in summary.itertuples(index=False):
        ws.append(row)

    # Format headers
    header_font = Font(name="Cambria", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Add chart
    chart = BarChart()
    chart.title = f"{period.capitalize()} Report"
    chart.x_axis.title = "Activity"
    chart.y_axis.title = "Total Hours"

    max_row = ws.max_row
    data = Reference(ws, min_col=3, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=2, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws._charts = []
    ws.add_chart(chart, "E2")

    wb.save(FILE_NAME)
    print(f"✅ {period.capitalize()} report exported successfully!")

# -----------------------
# MAIN FEATURES
# -----------------------
def add_entry():
    """Add new entry to Tracker + update Reports + Summary."""
    date_str = input("Enter date (YYYY-MM-DD) [empty for today]: ").strip()
    if date_str:
        try:
            date = dt.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            print("⚠ Invalid date, using today.")
            date = dt.date.today()
    else:
        date = dt.date.today()

    print("\nChoose activity:")
    for i, act in enumerate(ACTIVITIES, 1):
        print(f"{i}. {act}")
    try:
        choice = int(input("Enter number: ").strip())
        activity = ACTIVITIES[choice - 1]
    except:
        print("⚠ Invalid choice.")
        return

    hours = parse_hours(input("Enter hours (e.g., 1.5 hr, 30 min): ").strip())
    if hours <= 0:
        print("⚠ Invalid hours.")
        return

    mood = get_mood()

    new_data = pd.DataFrame({"Date": [date], "Task": [activity], "Hours": [hours], "Mood": [mood]})
    if os.path.exists(FILE_NAME):
        try:
            existing = pd.read_excel(FILE_NAME, sheet_name=SHEET_TRACKER)
            df = pd.concat([existing, new_data], ignore_index=True)
        except:
            df = new_data
    else:
        df = new_data

    with pd.ExcelWriter(FILE_NAME, engine="openpyxl",
                        mode="a" if os.path.exists(FILE_NAME) else "w",
                        if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=SHEET_TRACKER, index=False)

    setup_reports_sheet()
    add_to_reports(activity, hours)
    setup_summary_sheet()
    update_summary_sheet()

    print("✅ Entry saved!")

def view_data():
    if not os.path.exists(FILE_NAME):
        print("⚠ No data yet.")
        return
    try:
        df = pd.read_excel(FILE_NAME, sheet_name=SHEET_TRACKER)
        print(df)
    except Exception as e:
        print("⚠ Error:", e)

# -----------------------
# OPEN EXCEL
# -----------------------
def open_excel():
    """Open the Excel file with the default program (Excel)."""
    if not os.path.exists(FILE_NAME):
        print("⚠ Excel file not found.")
        return

    try:
        if sys.platform.startswith("win"):
            os.startfile(FILE_NAME)  # ✅ Windows
        elif sys.platform.startswith("darwin"):
            subprocess.call(["open", FILE_NAME])  # macOS
        else:
            subprocess.call(["xdg-open", FILE_NAME])  # Linux
        print("✅ Excel file opened.")
    except Exception as e:
        print("⚠ Could not open Excel:", e)

# -----------------------
# MAIN MENU
# -----------------------
def main():
    while True:
        print("\n====== Daily Life Tracker ======")
        print("1. Add new entry")
        print("2. View all data")
        print("3. Open Excel")
        print("4. Export report (daily/weekly/monthly)")
        print("5. Exit")

        choice = input("Enter choice: ").strip()
        if choice == "1":
            add_entry()
        elif choice == "2":
            view_data()
        elif choice == "3":
            open_excel()
        elif choice == "4":
            print("\nChoose report type:")
            print("1. Daily Report")
            print("2. Weekly Report")
            print("3. Monthly Report")
            sub_choice = input("Enter number: ").strip()
            if sub_choice == "1":
                export_report("daily")
            elif sub_choice == "2":
                export_report("weekly")
            elif sub_choice == "3":
                export_report("monthly")
            else:
                print("⚠ Invalid choice.")
        elif choice == "5":
            break
        else:
            print("⚠ Invalid choice.")


if __name__ == "__main__":
    ensure_sheets_exist()
    main()

